# import openpyxl module
import openpyxl

# Call a Workbook() function of openpyxl
# to create a new blank Workbook object
wb = openpyxl.Workbook()

# Get workbook active sheet
# from the active attribute.
sheet = wb.active

# Once have the Worksheet object,
# one can get its name from the
# title attribute.
sheet_title = sheet.title

print("active sheet title: " + sheet_title)
# One can change the name of the title
sheet.title = "sheet1"
  
print("sheet name is renamed as: " + sheet.title)

# Cell objects also have row, column
# and coordinate attributes that provide
# location information for the cell.
  
# Note: The first row or column integer
# is 1, not 0. Cell object is created by
# using sheet object's cell() method.
c1 = sheet.cell(row = 1, column = 1)
  
# writing values to cells
c1.value = "ANKIT"
  
c2 = sheet.cell(row= 1 , column = 2)
c2.value = "RAI"
  
# Once have a Worksheet object, one can
# access a cell object by its name also.
# A2 means column = 1 & row = 2.
c3 = sheet['A2']
c3.value = "RAHUL"
  
# B2 means column = 2 & row = 2.
c4 = sheet['B2']
c4.value = "RAI"
  
# Anytime you modify the Workbook object
# or its sheets and cells, the spreadsheet
# file will not be saved until you call
# the save() workbook method.
wb.save("demoWrite.xlsx")
print("----------------------------")

# Sheets can be added to workbook with the
# workbook object's create_sheet() method. 
wb.create_sheet(index = 1 , title = "demo sheet2")
  
wb.save("demoWrite.xlsx")








